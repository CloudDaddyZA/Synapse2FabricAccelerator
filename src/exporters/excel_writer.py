"""Excel workbook writer built on openpyxl with consultant-ready styling."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_workbook(sheets: dict[str, list[dict[str, Any]]], path: str | Path) -> Path:
    """Write multiple sheets. ``sheets`` maps sheet name -> list of row dicts."""
    dest = Path(path)
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            ws.append(["(no data)"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for col_idx, _header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([_stringify(row.get(h, "")) for h in headers])
        for col_idx, header in enumerate(headers, start=1):
            width = max(len(str(header)), *(len(_stringify(r.get(header, ""))) for r in rows))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 12), 60)
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        wb.create_sheet("Empty")
    wb.save(dest)
    return dest


def _stringify(value: Any) -> str:
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value)
    if hasattr(value, "value"):
        return str(value.value)
    return str(value)
